from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os


def write_name_on_certificate(name, certificate_template_path, output_path):
    # Load the certificate template image
    certificate_template = Image.open(certificate_template_path)
    draw = ImageDraw.Draw(certificate_template)

    # Specify the font path and size
    font_path = "/home/ghrce/Desktop/cert/ITCEDSCR.TTF"  # Replace with the actual font path
    font_size = 200	

    # Load the font
    font = ImageFont.truetype(font_path, font_size)

    # Calculate the position to write the name on the certificate
    text_width, text_height = draw.textsize(name, font=font)
    position = ((certificate_template.width - text_width) / 2, (certificate_template.height - text_height) / 2.5)

    # Write the name on the certificate
    draw.text(position, name, fill="black", font=font)

    # Save the certificate as a JPEG file
    certificate_template.save(output_path, "JPEG")


def excel_to_certificates(excel_file, certificate_template, output_directory):
    # Load the Excel file
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Iterate through the names in the Excel sheet
    for row in sheet.iter_rows(values_only=True):
        name = row[0]  # Assuming the name is in the first column of the Excel sheet

        # Generate the output file path
        output_file = f"{output_directory}/{name}.jpeg"

        # Generate the certificate with the name
        write_name_on_certificate(name, certificate_template, output_file)

        # Rename the file
        new_output_file = f"{output_directory}/{row[1]}.jpeg"  # Assuming the new name is in the second column of the Excel sheet
        os.rename(output_file, new_output_file)

        print(f"Generated and renamed certificate for {name}.")


print("All certificates generated and renamed successfully.")


# Specify the paths to the Excel file, certificate template image, and output directory
excel_file_path = "/home/ghrce/Desktop/cert/students.xlsx"
certificate_template_path = "/home/ghrce/Desktop/cert/Participants.jpg"
output_directory_path = "/home/ghrce/Desktop/print"

# Generate certificates and rename the files
excel_to_certificates(excel_file_path, certificate_template_path, output_directory_path)
